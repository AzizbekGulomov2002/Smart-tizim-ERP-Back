from django.core.exceptions import ObjectDoesNotExist
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, status, filters, generics
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
from openpyxl import load_workbook
from django.db import transaction
from apps.app.views import BasePagination
from apps.products.filters import ProductFilter, FormatFilter, CategoryFilter, StorageFilter
from apps.products.serializers import  *
from apps.products.decorator import is_storage_permission, is_product_permission
from datetime import datetime, timedelta
from rest_framework.views import APIView
import json
from apps.finance.models import Transaction, FinanceOutcome


# Product Delete Manager
class ProductDeleteManagerAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        company_id = self.request.user.company_id
        all_delete = datetime.now().date() - timedelta(days=30) # Auto delete in 30 days
        delete_products = Product.all_objects.filter(company_id=company_id,deleted__lt=all_delete)
        if delete_products.exists():
            delete_products.delete()
        deleted = datetime.now().date() - timedelta(days=30)
        clients = Product.all_objects.filter(company_id=company_id,deleted__gte=deleted)
        serializer = ProductSerializer(clients, many=True)
        return Response(serializer.data)
    @is_product_permission
    def post(self,request):
        data = request.data
        try:
            id =  data['product']['id']
            product = Product.all_objects.get(id=id)
            product.restore()
        except: pass
        return Response({"status":'ok'})
    @is_product_permission
    def delete(self,request):
        company_id = self.request.user.company_id
        # print(request.user)
        # print(company_id)
        all_delete = datetime.now().date()
        client =  Product.all_objects.filter(company_id=company_id,deleted__lte=all_delete)
        # print(company_id)
        client.delete()
        return Response({"status":'ok'})


class AllFormatViewSet(ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FormatSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_class = FormatFilter
    ordering_fields = ['name']
    search_fields = ['name']

    def get_queryset(self):
        company_id = self.request.user.company_id
        queryset = Format.objects.filter(company_id=company_id)
        return queryset


class FormatViewSet(ModelViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = BasePagination
    serializer_class = FormatSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_class = FormatFilter
    ordering_fields = ['name']
    search_fields = ['name']

    def get_queryset(self):
        company_id = self.request.user.company_id
        queryset = Format.objects.filter(company_id=company_id)
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(company_id=request.user.company_id)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CategoryViewSet(ModelViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = BasePagination
    serializer_class = CategorySerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_class = CategoryFilter  # Assuming you have defined CategoryFilter
    ordering_fields = ['name']
    search_fields = ['name']

    def get_queryset(self):
        # company_id = 1
        company_id = self.request.user.company_id
        queryset = Category.objects.filter(company_id=company_id)
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(company_id=request.user.company_id)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

class AllCategoryViewSet(ModelViewSet):
    permission_classes= [IsAuthenticated]
    serializer_class = CategorySerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_class = CategoryFilter
    ordering_fields = ['name']
    search_fields = ['name']
    def get_queryset(self):
        company_id = self.request.user.company_id
        queryset = Category.objects.filter(company_id=company_id)
        return queryset


class ProductViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = ProductSerializer
    pagination_class = BasePagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = ProductFilter
    search_fields = ("category__name", "format__name", "name", "bar_code", "price")

    def get_queryset(self):
        company_id = self.request.user.company_id
        queryset = Product.objects.filter(company_id=company_id).order_by('-id')
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        return Response(status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, partial=True)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ProductImportView(APIView):
    def post(self, request, *args, **kwargs):
        company_id = request.user.company_id
        serializer = ProductImportSerializer(data=request.data)

        if serializer.is_valid():
            file = serializer.validated_data['file']
            try:
                workbook = load_workbook(file)
                sheet = workbook.active
                imported_products = []
                with transaction.atomic():
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        category_name = row[0]
                        format_name = row[1]
                        product_name = row[2]
                        product_type = row[3]
                        price = int(row[4])
                        bar_code = row[5] if row[5] else None
                        storage_name = row[6] if row[6] else None  # Assuming storage name is in column 7

                        # Validate product_type and storage_name
                        if product_type == "Sanaladigan" and not storage_name:
                            raise ValueError(f"'{product_name}' nomli mahsulot Omborda 'Sanaladigan' turda. Unga ombor nomini kiritish kerak !")
                        elif product_type == "Sanalmaydigan" and storage_name:
                            storage_name = None  # Ignore storage_name for 'Sanalmaydigan' products

                        # Check if category exists
                        category, category_created = Category.objects.get_or_create(name=category_name, company_id=company_id)
                        if category_created:
                            raise ValueError(f"Kategoriya '{category_name}' mavjud emas")

                        # Check if format exists
                        format, format_created = Format.objects.get_or_create(name=format_name, company_id=company_id)
                        if format_created:
                            raise ValueError(f"Format '{format_name}' mavjud emas")

                        # Check if storage exists
                        if storage_name:
                            storage, storage_created = Storage.objects.get_or_create(name=storage_name, company_id=company_id)
                            if storage_created:
                                raise ValueError(f"'{storage_name}' korxonada ombor mavjud emas ")
                        else:
                            storage = None

                        product = Product.objects.create(
                            company_id=company_id,
                            name=product_name,
                            product_type=product_type,
                            category=category,
                            format=format,
                            price=price,
                            bar_code=bar_code,
                            storage=storage
                        )

                        imported_products.append(ProductSerializer(product).data)

                return Response({
                    "success": "Mahsulotlar mucaffaqiyatli import qilindi",
                    "products": imported_products
                }, status=status.HTTP_201_CREATED)

            except ValueError as ve:
                return Response({"error": str(ve)}, status=status.HTTP_400_BAD_REQUEST)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class ALlProductViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = ProductSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = ProductFilter
    search_fields = ("protype__name", "action_type", "storage_date", "storage_count")

    def get_queryset(self):
        company_id = self.request.user.company_id
        queryset = Product.objects.filter(company_id=company_id).order_by('-id')
        return queryset


class ProductCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company_id = request.user.company_id
        products = Product.objects.filter(company_id=company_id)
        categories = Category.objects.filter(company_id=company_id)
        formats = Format.objects.filter(company_id=company_id)
        storages = Storage.objects.all()  # Fetch all storages

        product_serializer = ProductSerializer(products, many=True)
        category_serializer = CategorySerializer(categories, many=True)
        format_serializer = FormatSerializer(formats, many=True)
        storage_serializer = StorageSerializer(storages, many=True)

        return Response({
            "products": product_serializer.data,
            "categories": category_serializer.data,
            "formats": format_serializer.data,
            "storages": storage_serializer.data  # Include storage data
        })

    def post(self, request):
        company_id = request.user.company_id
        data = request.data

        # Ensure data is treated as a list for consistent processing
        if isinstance(data, dict):
            data = [data]

        created_products = {}
        errors = {}

        for item in data:
            name = item.get("name")
            price = item.get("price")
            category_id = item.get("category_id")
            format_id = item.get("format_id")
            product_type = item.get("product_type")
            bar_code = item.get("bar_code", "")
            storage_id = item.get("storage_id")  # Get storage_id from request data

            if not all([name, price, category_id, format_id, storage_id]):
                errors.update({"error": "name, price, category_id, format_id, and storage_id are required for each product."})
                continue

            try:
                category = Category.objects.get(company_id=company_id, id=category_id)
            except Category.DoesNotExist:
                errors.update({"error": f"Invalid category_id {category_id} for the given company."})
                continue

            try:
                format = Format.objects.get(company_id=company_id, id=format_id)
            except Format.DoesNotExist:
                errors.update({"error": f"Invalid format_id {format_id} for the given company."})
                continue

            try:
                storage = Storage.objects.get(id=storage_id)
            except Storage.DoesNotExist:
                errors.update({"error": f"Invalid storage_id {storage_id}."})
                continue

            # Create the product
            product = Product(
                name=name,
                company_id=company_id,
                product_type=product_type,
                category=category,
                format=format,
                price=price,
                bar_code=bar_code,
                storage=storage  # Assign storage to the product
            )
            product.save()
            created_products.update({product.id: ProductSerializer(product).data})

        if errors:
            return Response({
                "message": "Errors encountered while processing your request.",
                "details": errors
            }, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "message": "Products created successfully.",
            "products": created_products
        }, status=status.HTTP_201_CREATED)

class SupplierViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = SupplierSerializer
    def get_queryset(self):
        company_id = self.request.user.company_id
        queryset = Supplier.objects.filter(company_id=company_id).order_by('-id')
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(company_id=request.user.company_id)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class AllStorageViewSet(ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = StorageSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    ordering_fields = ['name']
    search_fields = ['name']
    def get_queryset(self):
        company_id = self.request.user.company_id
        queryset = Storage.objects.filter(company_id=company_id)
        return queryset


class StorageViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = StorageSerializer
    pagination_class = BasePagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = StorageFilter
    search_fields = ("name")
    def get_queryset(self):
        company_id = self.request.user.company_id
        queryset = Storage.objects.filter(company_id=company_id)
        return queryset
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    @is_storage_permission
    def retrieve(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    @is_storage_permission
    # limit
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(company_id=request.user.company_id)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @is_storage_permission
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @is_storage_permission
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

class StorageProductViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = BasePagination
    serializer_class = StorageProductSerializer
    def get_queryset(self):
        company_id = self.request.user.company_id
        queryset = StorageProduct.objects.filter(company_id=company_id).order_by('-id')
        return queryset
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @is_storage_permission
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @is_storage_permission
    def create(self, request, *args, **kwargs):
        # serialaizer = self.get_serializer(data=request.data)
        # if serialaizer.is_valid():
        #     serialaizer.save(company_id=request.user.company_id)
        #     return Response(serialaizer.data, status=status.HTTP_201_CREATED)
        return Response(status=status.HTTP_400_BAD_REQUEST)

    @is_storage_permission
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @is_storage_permission
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


def parse_datetime(date_str = None):
    try: return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
    except:return None

def parse_date(date_str = None):
    try:return datetime.strptime(date_str, "%Y-%m-%d").date()
    except:return None


class StorageProductCreate(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        company_id = request.user.company_id
        products = Product.objects.select_related('format', 'category').prefetch_related('storage_products').filter(
            company_id=company_id)
        categories = Category.objects.filter(company_id=company_id)
        suppliers = Supplier.objects.filter(company_id=company_id)
        formats = Format.objects.filter(company_id=company_id)

        product_serializer = ProductSerializer(products, many=True)
        category_serializer = CategorySerializer(categories, many=True)
        supplier_serializer = SupplierSerializer(suppliers, many=True)
        format_serializer = FormatSerializer(formats, many=True)

        return Response({
            "supplier": supplier_serializer.data,
            "category": category_serializer.data,
            "format": format_serializer.data,
            "product": product_serializer.data
        })

    def post(self, request):
        company_id = request.user.company_id
        data = json.loads(request.body.decode('utf-8'))
        supplier = None
        storage = None
        finance_outcome_data = None

        try:
            for item in data:
                if "supplier" in item:
                    supplier_id = item.get("supplier")
                    supplier = Supplier.objects.get(id=supplier_id, company_id=company_id)
                    break

            for item in data:
                if "storage" in item:
                    storage_id = item.get("storage")
                    storage = Storage.objects.get(id=storage_id, company_id=company_id)
                    break

            for item in data:
                if "cash" in item or "card" in item or "other_pay" in item or "desc" in item:
                    cash= item.get("cash", 0)
                    card= item.get("card", 0)
                    other_pay= item.get("other_pay", 0)
                    total = cash+card+other_pay
                    finance_outcome_data = {
                        "user": request.user,
                        "supplier": supplier,
                        "cash": item.get("cash", 0),
                        "card": item.get("card", 0),
                        "other_pay": item.get("other_pay", 0),
                        "total": item.get("total",total),
                        "desc": item.get("desc", ""),
                        "deadline":item.get("deadline"),
                        "company_id": company_id
                    }
                    break

            if not supplier:
                return Response({"error": "Supplier not found"}, status=status.HTTP_400_BAD_REQUEST)

            if not storage:
                return Response({"error": "Storage not found"}, status=status.HTTP_400_BAD_REQUEST)

            storage_products = []
            for item in data:
                if "supplier" in item or "storage" in item or "cash" in item or "card" in item or "other_pay" in item or "desc" in item:
                    continue

                product_id = item.get("product")
                if product_id:
                    product = Product.objects.get(id=product_id, company_id=company_id)
                else:
                    # Handle creation of new product if product_id is not provided
                    name = item.get("name")
                    price = item.get("price")
                    category_id = item.get("category")
                    format_id = item.get("format")
                    product, created = Product.objects.get_or_create(
                        name=name,
                        company_id=company_id,
                        defaults={
                            'product_type': item.get("product_type"),
                            'category_id': category_id,
                            'format_id': format_id,
                            'price': price,
                        }
                    )
                    if not created and product.price != price:
                        product.price = price
                        product.save()

                # Parse date and expiration fields
                date_str = item.get("date")
                if not date_str:
                    return Response({"error": "Date is required"}, status=status.HTTP_400_BAD_REQUEST)
                try:
                    date = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S")
                except ValueError:
                    return Response({"error": f"Invalid date format: {date_str}"},
                                    status=status.HTTP_400_BAD_REQUEST)

                expiration = parse_date(item.get("expiration"))
                remind_count = item.get("remind_count") or 0
                size_type = item.get("size_type") or "O'lchovsiz"
                storage_type = item.get("storage_type") or "Naqtga"
                storage_price = item.get("price")

                if storage_price is None:
                    return Response({"error": "Price is required"}, status=status.HTTP_400_BAD_REQUEST)

                storage_product = StorageProduct(
                    company_id=company_id,
                    storage_type=storage_type,
                    size_type=size_type,
                    storage_count=item.get("storage_count"),
                    part_size=item.get("part_size"),
                    height=item.get("height"),
                    width=item.get("width"),
                    product=product,
                    price=storage_price,
                    date=date,
                    user=request.user,
                    supplier=supplier,
                    storage=storage,
                    remind_count=remind_count,
                    expiration=expiration,
                )
                storage_product.save()
                storage_products.append(storage_product)

            try:
                if finance_outcome_data:
                    transaction_type = Transaction.objects.get(transaction_type='ombor', action_type='chiqim')
                    finance_outcome_data["tranzaction_type"] = transaction_type

                    # Create one FinanceOutcome for all StorageProducts
                    finance_outcome = FinanceOutcome.objects.create(**finance_outcome_data)
                    # Associate the FinanceOutcome with each StorageProduct
                    total=0
                    for storage_product in storage_products:
                        total+=(storage_product.price*storage_product.storage_count)
                        storage_product.finance_outcome = finance_outcome
                        storage_product.save()
                    finance_outcome.total=total
                    finance_outcome.save()

            except ObjectDoesNotExist:
                return Response({"error": "Transaction type 'ombor' with action 'chiqim' does not exist"},
                                status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({"error": str(e.args)},status=status.HTTP_400_BAD_REQUEST)

            return Response({"status": "ok"}, status=status.HTTP_201_CREATED)

        except Supplier.DoesNotExist:
            return Response({"error": "Supplier not found"}, status=status.HTTP_400_BAD_REQUEST)
        except Storage.DoesNotExist:
            return Response({"error": "Storage not found"}, status=status.HTTP_400_BAD_REQUEST)
        except Product.DoesNotExist:
            return Response({"error": "Product not found"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class StorageProductOffViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = StorageProductOffSerializer

    def get_queryset(self):
        company_id = self.request.user.company_id
        return StorageProductOff.objects.filter(product__company_id=company_id)

    # def create(self, request, *args, **kwargs):
    #     serializer = self.get_serializer(data=request.data)
    #     if serializer.is_valid():
    #         serializer.save()
    #         return Response(serializer.data, status=status.HTTP_201_CREATED)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def create(self, request, *args, **kwargs):
        serializers = self.get_serializer(data=request.data)
        if serializers.is_valid():
            serializers.save(company_id=request.user.company_id)
            return Response(serializers.data, status=status.HTTP_201_CREATED)
        return Response(serializers.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class StorageProductTransferViewSet(viewsets.ModelViewSet):
    queryset = StorageProductTransfer.objects.all()
    serializer_class = StorageProductTransferSerializer
    permission_classes = [IsAuthenticated]